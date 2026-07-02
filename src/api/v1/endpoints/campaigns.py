import os
import re

import pandas as pd
import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from typing import Any, List, Dict
from datetime import datetime, timedelta
from io import BytesIO
from celery import group

from fastapi import APIRouter, Body, Depends, Request, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder

from sqlalchemy.ext.asyncio import AsyncSession

from core.config import settings
from api import deps
from tasks import prepare_campaign, check_dst_batch

import crud, models, schemas
from services.android import AndroidService

router = APIRouter()


def extract_links(msg_template: str, dst_data: Dict) -> List[str]:
    """
    Extract links from msg_template that need to be shortened.

    Args:
        msg_template: Template with [short]...[/short] placeholders
        dst_data: Dictionary with field values (field_1, field_2, etc.)

    Returns:
        List of unique URLs to shorten
    """
    pattern = r"\[short\](.+?)\[/short\]"
    matches = re.findall(pattern, msg_template)

    urls = []
    for match in matches:
        # Check if it's a field reference like {field_1}
        if match.startswith("{field_") and match.endswith("}"):
            field_name = match[1:-1]  # Remove { }
            field_value = dst_data.get(field_name, "")
            if field_value:
                urls.append(field_value)
        else:
            # Direct URL in template
            urls.append(match)

    # Return unique URLs preserving order
    seen = set()
    unique_urls = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            unique_urls.append(url)

    return unique_urls


@router.get("/", response_model=schemas.CampaignRows)
async def read_campaigns(
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    filters: List[schemas.Filter] = Depends(deps.request_filters),
    orders: List[schemas.Order] = Depends(deps.request_orders),
    skip: int = 0,
    limit: int = 100,
) -> Any:
    """
    Retrieve campaigns.
    """
    if not orders:
        orders = [{"field": "id", "dir": "desc"}]
    filters_modified = []
    tags_idx = None
    status_idx = None

    for filter in filters:
        if sub_filters := filter.get("filters"):
            for filter in sub_filters:
                filters_modified.append(filter)
        else:
            filters_modified.append(filter)
    filters = []
    for i in range(len(filters_modified)):
        if filters_modified[i]["field"] == "tags":
            if tags_idx is None:
                tags_idx = len(filters)
                filters.append(
                    {
                        "relationship": models.Campaign.campaign_tags,
                        "field": models.CampaignTags.tag_id,
                        "operator": "overlaps",
                        "value": [filters_modified[i]["value"]],
                    }
                )
            else:
                filters[tags_idx]["value"].append(filters_modified[i]["value"])
        elif filters_modified[i]["field"] == "status":
            if status_idx is None:
                status_idx = len(filters)
                filters.append(
                    {
                        "field": "status",
                        "operator": "overlaps",
                        "value": [filters_modified[i]["value"]],
                    }
                )
            else:
                filters[status_idx]["value"].append(filters_modified[i]["value"])
        else:
            filters.append(filters_modified[i])

    if current_user.is_superuser:
        campaigns = await crud.campaign.get_rows(
            db, skip=skip, limit=limit, filters=filters, orders=orders
        )
        count = await crud.campaign.get_count(db, filters=filters)
    else:
        campaigns = await crud.campaign.get_rows_by_user(
            db,
            user_id=current_user.id,
            filters=filters,
            orders=orders,
            skip=skip,
            limit=limit,
        )
        count = await crud.campaign.get_count_by_user(
            db, user_id=current_user.id, filters=filters
        )

    # Получаем данные Android устройств
    android = AndroidService()
    data = await android.get_device_options(x_api_key=current_user.ext_api_key)

    # Создаём словарь для быстрого маппинга value -> text
    android_map = {
        item["value"]: item["text"]
        for item in data
        if "value" in item and "text" in item
    }

    # Обогащаем каждую кампанию полем android_names
    for campaign in campaigns:
        campaign.android_names = []
        if campaign.androids:
            for android_id in campaign.androids:
                android_name = android_map.get(android_id, android_id)
                campaign.android_names.append(android_name)

    return {"data": campaigns, "total": count}


@router.post("/", response_model=schemas.Campaign)
async def create_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    campaign_in: schemas.CampaignRequest,
) -> Any:
    """
    Create new campaign.
    """

    ts = datetime.utcnow()
    campaign_user_id = campaign_in.user_id if campaign_in.user_id else current_user.id
    can_assign_api_keys = (
        current_user.is_superuser
        or "campaign.assign_api_keys" in current_user.permissions
    )
    campaign_api_keys = campaign_in.api_keys

    if not can_assign_api_keys:
        campaign_user = current_user
        if campaign_user.id != campaign_user_id:
            campaign_user = await crud.user.get(db=db, id=campaign_user_id)
        campaign_api_keys = list(campaign_user.api_keys or []) if campaign_user else []

    campaign_db_in = schemas.CampaignCreate(
        name=campaign_in.name,
        user_id=campaign_user_id,
        webhook_url=campaign_in.webhook_url,
        api_keys=campaign_api_keys,
        androids=campaign_in.androids,
        tags=campaign_in.tags,
        schedule=campaign_in.schedule,
        msg_attempts=campaign_in.msg_attempts,
        msg_sending_timeout=campaign_in.msg_sending_timeout,
        msg_status_timeout=campaign_in.msg_status_timeout,
        msg_template=campaign_in.msg_template,
        follow_limit=campaign_in.follow_limit,
        order=campaign_in.order,
        create_ts=ts,
        start_ts=campaign_in.start_ts,
        stop_ts=campaign_in.stop_ts,
        status=campaign_in.status,
    )

    campaign_dst_in = []
    fields = campaign_in.data_fields

    if (
        campaign_in.data_text
        and campaign_in.data_text_row_sep
        and campaign_in.data_text_col_sep
    ):
        for line in campaign_in.data_text.split(campaign_in.data_text_row_sep):
            campaign_dst = {}
            campaign_dst["attempts"] = campaign_in.msg_attempts
            if campaign_in.msg_sending_timeout:
                campaign_dst["expire_ts"] = ts + timedelta(
                    seconds=campaign_in.msg_sending_timeout
                )
            row = line.split(campaign_in.data_text_col_sep)
            for f in fields:
                if (
                    f
                    in [
                        "dst_addr",
                        "field_1",
                        "field_2",
                        "field_3",
                        "field_4",
                        "field_5",
                    ]
                    and fields[f] is not None
                    and int(fields[f]) < len(row)
                ):
                    campaign_dst[f] = row[int(fields[f])]
            campaign_dst_in.append(campaign_dst)

    if campaign_in.data_file_name:
        ext = os.path.splitext(campaign_in.data_file_name)[1]
        data = pd.DataFrame()
        if ext in [".csv", ".txt"]:
            try:
                data = pd.read_csv(
                    f"upload/{campaign_in.data_file_name}", sep="[\s+|,;:]", header=None
                )
            except pd.errors.ParserError:
                data = pd.read_csv(f"upload/{campaign_in.data_file_name}", header=None)
            data = data.astype(str)
        if ext in [".xls", ".xlsx"]:
            data = pd.read_excel(
                f"upload/{campaign_in.data_file_name}", sheet_name=0, header=None
            )
            data = data.astype(str)
        for i, row in data.iterrows():
            campaign_dst = {}
            campaign_dst["attempts"] = campaign_in.msg_attempts
            if campaign_in.msg_sending_timeout:
                campaign_dst["expire_ts"] = ts + timedelta(
                    seconds=campaign_in.msg_sending_timeout
                )
            for i in fields:
                if (
                    i
                    in [
                        "dst_addr",
                        "field_1",
                        "field_2",
                        "field_3",
                        "field_4",
                        "field_5",
                    ]
                    and fields[i] is not None
                    and int(fields[i]) in row
                ):
                    campaign_dst[i] = row[int(fields[i])].strip().strip('"').strip("'")
            campaign_dst_in.append(campaign_dst)

    campaign_db_in.msg_total = len(campaign_dst_in)
    campaign = await crud.campaign.create(db=db, obj_in=campaign_db_in)

    if campaign.id and len(campaign_dst_in) > 0:
        # Добавляем campaign_id и заполняем text с подстановкой полей для всех записей
        for dst_data in campaign_dst_in:
            dst_data["campaign_id"] = campaign.id

            # Подстановка полей {field_1} ... {field_5}
            text_with_fields = campaign.msg_template or ""
            for field_name in [
                "field_1",
                "field_2",
                "field_3",
                "field_4",
                "field_5",
            ]:
                if dst_data.get(field_name):
                    text_with_fields = text_with_fields.replace(
                        "{" + field_name + "}", dst_data[field_name]
                    )
            dst_data["text"] = text_with_fields

        # Проверяем наличие ссылок для сокращения
        has_links_to_shorten = False
        if campaign.msg_template:
            for dst_data in campaign_dst_in:
                links = extract_links(dst_data.get("text", campaign.msg_template), dst_data)
                if links:
                    has_links_to_shorten = True
                    break

        # Определяем шаги для prepare_campaign (сокращение ссылок и/или рерайт)
        steps = []
        if has_links_to_shorten:
            steps.append("shorten_links")
        if campaign_in.rewrite and campaign_in.provider:
            steps.append("rewrite")

        # Создаем записи campaign_dst
        # Статус WAITING только если есть сокращение ссылок или рерайт
        if steps:
            for dst_data in campaign_dst_in:
                dst_data["status"] = schemas.CampaignDstStatus.WAITING

            created_objects = await crud.campaign_dst.create_rows(
                db=db, obj_in=campaign_dst_in
            )

            # Запускаем Celery задачу предобработки (ссылки + рерайт)
            rewrite_config = None
            if campaign_in.rewrite and campaign_in.provider:
                api_key = ""
                if campaign_in.provider.lower() == "ollama":
                    api_key = settings.AI_OLLAMA_API_KEY
                elif campaign_in.provider.lower() == "openrouter":
                    api_key = settings.AI_OPENROUTER_API_KEY

                rewrite_config = {
                    "provider": campaign_in.provider,
                    "model": campaign_in.model,
                    "api_key": api_key,
                    "prompt": campaign_in.prompt or settings.AI_REWRITE_SYSTEM_PROMPT,
                }

            prepare_campaign.delay(
                campaign_id=campaign.id,
                steps=steps,
                x_api_key=current_user.ext_api_key,
                rewrite_config=rewrite_config,
            )
        else:
            # Нет сокращения ссылок или рерайта - сразу статус CREATED
            for dst_data in campaign_dst_in:
                dst_data["status"] = schemas.CampaignDstStatus.CREATED

            created_objects = await crud.campaign_dst.create_rows(
                db=db, obj_in=campaign_dst_in
            )

        # НЕЗАВИСИМАЯ проверка номеров (если требуется)
        # Не влияет на статус campaign_dst, запускается параллельно
        if campaign_in.check_dst:
            # Создаем батчи для проверки
            check_batches = [
                created_objects[i: i + settings.PHONE_CHECKER_BATCH_SIZE]
                for i in range(0, len(created_objects), settings.PHONE_CHECKER_BATCH_SIZE)
            ]

            check_tasks = []
            for batch in check_batches:
                batch_data = [
                    {"id": dst["id"], "phone": dst["dst_addr"]}
                    for dst in batch if dst.get("dst_addr")
                ]
                if batch_data:
                    check_tasks.append(check_dst_batch.si(batch_data))

            if check_tasks:
                job = group(*check_tasks)
                job.apply_async()

    return campaign


@router.post("/start", response_model=List[schemas.Campaign])
async def update_campaign_rows(
    *,
    current_user: models.User = Depends(deps.get_current_active_user),
    db: AsyncSession = Depends(deps.get_db),
    ids: List[int] = Body(..., embed=True),
) -> Any:
    """
    Start a campaigns.
    """
    user_id = None if current_user.is_superuser else current_user.id
    result = await crud.campaign.update_rows(
        db, ids=ids, user_id=user_id, obj_in={"status": schemas.CampaignStatus.RUNNING}
    )
    return result


@router.post("/stop", response_model=List[schemas.Campaign])
async def update_campaign_rows(
    *,
    current_user: models.User = Depends(deps.get_current_active_user),
    db: AsyncSession = Depends(deps.get_db),
    ids: List[int] = Body(..., embed=True),
) -> Any:
    """
    Start a campaigns.
    """
    user_id = None if current_user.is_superuser else current_user.id
    result = await crud.campaign.update_rows(
        db, ids=ids, user_id=user_id, obj_in={"status": schemas.CampaignStatus.STOPPED}
    )
    return result


@router.delete("/", response_model=List[int])
async def delete_campaign_rows(
    *,
    current_user: models.User = Depends(deps.get_current_active_user),
    db: AsyncSession = Depends(deps.get_db),
    ids: List[int] = Body(..., embed=True),
) -> Any:
    """
    Start a campaigns.
    """
    user_id = None if current_user.is_superuser else current_user.id
    result = await crud.campaign.delete_rows(db, ids=ids, user_id=user_id)
    return result


@router.put("/{id}", response_model=schemas.Campaign)
async def update_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    _current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
    campaign_in: schemas.CampaignUpdate,
) -> Any:
    """
    Update a campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    campaign = await crud.campaign.update(db=db, db_obj=campaign, obj_in=campaign_in)
    if campaign.create_ts:
        campaign.create_ts = campaign.create_ts.strftime("%Y-%m-%d %H:%M:%S")
    if campaign.start_ts:
        campaign.start_ts = campaign.start_ts.strftime("%Y-%m-%d %H:%M:%S")
    if campaign.stop_ts:
        campaign.stop_ts = campaign.stop_ts.strftime("%Y-%m-%d %H:%M:%S")
    return campaign


@router.post("/{id}/start", response_model=schemas.Campaign)
async def update_campaign(
    *,
    current_user: models.User = Depends(deps.get_current_active_user),
    db: AsyncSession = Depends(deps.get_db),
    id: int,
) -> Any:
    """
    Start a campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    start_ts = campaign.start_ts if campaign.start_ts else datetime.utcnow()
    campaign = await crud.campaign.update(
        db=db,
        db_obj=campaign,
        obj_in={"status": schemas.CampaignStatus.RUNNING, "start_ts": start_ts},
    )
    return campaign


@router.post("/{id}/pause", response_model=schemas.Campaign)
async def update_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
) -> Any:
    """
    Pause an campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    stop_ts = campaign.stop_ts if campaign.stop_ts else datetime.utcnow()
    campaign = await crud.campaign.update(
        db=db, db_obj=campaign, obj_in={"status": schemas.CampaignStatus.PAUSED}
    )
    return campaign


@router.post("/{id}/stop", response_model=schemas.Campaign)
async def update_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
) -> Any:
    """
    Stop an campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    stop_ts = campaign.stop_ts if campaign.stop_ts else datetime.utcnow()
    campaign = await crud.campaign.update(
        db=db,
        db_obj=campaign,
        obj_in={"status": schemas.CampaignStatus.STOPPED, "stop_ts": stop_ts},
    )
    return campaign


@router.post("/{id}/check_dst", response_model=schemas.Campaign)
async def check_campaign_dst(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
) -> Any:
    """
    Checking for all numbers in campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # Fetch all campaign_dst records for this campaign
    filters = [{"field": "campaign_id", "operator": "eq", "value": id}]
    campaign_dsts = await crud.campaign_dst.get_all(db=db, filters=filters)

    if not campaign_dsts:
        raise HTTPException(
            status_code=404, detail="No phone numbers found in campaign"
        )

    # Создаем батчи для проверки
    check_batches = [
        campaign_dsts[i : i + settings.PHONE_CHECKER_BATCH_SIZE]
        for i in range(0, len(campaign_dsts), settings.PHONE_CHECKER_BATCH_SIZE)
    ]

    check_tasks = []
    for batch in check_batches:
        batch_data = [
            {"id": dst.id, "phone": dst.dst_addr}
            for dst in batch if dst.dst_addr
        ]
        if batch_data:
            check_tasks.append(check_dst_batch.si(batch_data))

    if check_tasks:
        job = group(*check_tasks)
        job.apply_async()

    return campaign


@router.get("/{id}", response_model=schemas.Campaign)
async def read_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
) -> Any:
    """
    Get campaign by ID.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return campaign


@router.delete("/{id}", response_model=schemas.Campaign)
async def delete_campaign(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
) -> Any:
    """
    Delete an campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    campaign = await crud.campaign.delete(db=db, id=id)
    return campaign


@router.get("/{id}/campaign_dst", response_model=schemas.CampaignDstRows)
async def read_campaign_campaign_dsts(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
    filters: List[schemas.Filter] = Depends(deps.request_filters),
    orders: List[schemas.Order] = Depends(deps.request_orders),
    skip: int = 0,
    limit: int = 100,
    request: Request = None,
) -> Any:
    """
    Get campaign campaign_dsts.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if not orders:
        orders = [{"field": "id", "dir": "asc"}]
    filters.append({"field": "campaign_id", "operator": "eq", "value": id})
    campaign_dsts = jsonable_encoder(
        await crud.campaign_dst.get_rows(
            db=db, skip=skip, limit=limit, filters=filters, orders=orders
        )
    )
    for i in range(len(campaign_dsts)):
        if not campaign_dsts[i]["text"]:
            campaign_dsts[i]["text"] = campaign.msg_template or ""
            for j in range(1, 6):
                field = f"field_{j}"
                if campaign_dsts[i][field]:
                    campaign_dsts[i]["text"] = campaign_dsts[i]["text"].replace(
                        "{" + field + "}", campaign_dsts[i][field]
                    )
    count = await crud.campaign_dst.get_count(db=db, filters=filters)
    return JSONResponse({"data": campaign_dsts, "total": count})


@router.delete("/{id}/campaign_dst", response_model=schemas.Campaign)
async def delete_campaign_campaign_dsts(
    *,
    db: AsyncSession = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_active_user),
    id: int,
    ids: List[int] = Body(None, embed=True),
) -> Any:
    """
    Delete campaign DSTs. If ids is provided, delete only those rows.
    Otherwise delete all DSTs for the campaign.
    """
    campaign = await crud.campaign.get(db=db, id=id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.user_id != current_user.id and not current_user.is_superuser:
        raise HTTPException(
            status_code=400, detail="The user doesn't have enough privileges"
        )
    user_id = None if current_user.is_superuser else current_user.id
    rowcount = await crud.campaign_dst.delete_rows(
        db, campaign_id=id, ids=ids, user_id=user_id
    )
    is_partial_delete = ids is not None
    new_msg_total = max(0, campaign.msg_total - rowcount) if is_partial_delete else 0
    campaign = await crud.campaign.update(
        db=db,
        db_obj=campaign,
        obj_in={
            "msg_total": new_msg_total,
            "msg_sent": campaign.msg_sent if is_partial_delete else 0,
            "msg_delivered": campaign.msg_delivered if is_partial_delete else 0,
            "msg_undelivered": campaign.msg_undelivered if is_partial_delete else 0,
            "msg_failed": campaign.msg_failed if is_partial_delete else 0,
        },
    )
    return campaign


@router.get("/{campaign_id}/report")
async def download_campaign_report(
    *, db: AsyncSession = Depends(deps.get_db), campaign_id: int
):
    campaign = await crud.campaign.get(db=db, id=campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    filters = [{"field": "campaign_id", "operator": "eq", "value": campaign_id}]
    orders = [{"field": "id", "dir": "asc"}]
    campaign_dsts = await crud.campaign_dst.get_rows(
        db=db, limit=campaign.msg_total, filters=filters, orders=orders
    )
    if not campaign_dsts:
        raise HTTPException(status_code=404, detail="No campaign messages found")

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Campaign_{campaign_id}_report"

    headers = [
        "Название рассылки",
        "Номер телефона",
        "Время отправки",
        "Статус",
        "Сообщение",
    ]
    sheet.append(headers)

    for dst in campaign_dsts:
        text = campaign.msg_template
        for j in range(1, 6):
            field = f"field_{j}"
            if getattr(dst, field):
                text = text.replace("{" + field + "}", getattr(dst, field))
        sheet.append(
            [
                campaign.name,
                dst.dst_addr,
                dst.sent_ts.strftime("%Y-%m-%d %H:%M:%S") if dst.sent_ts else "",
                schemas.CampaignDstStatus.name(dst.status),
                ILLEGAL_CHARACTERS_RE.sub("", (dst.text or text)),
            ]
        )

    table_ref = "A1:{}".format(
        sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate
    )
    table = Table(displayName="DataTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    for column in sheet.columns:
        adjusted_width = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width + 5

    workbook.save(output)
    output.seek(0)

    filename = "{}_report_{}.xlsx".format(
        campaign_id, datetime.utcnow().strftime("%Y%m%d%H%M%S")
    )
    headers = {"Content-Disposition": f"attachment; filename={filename}"}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
